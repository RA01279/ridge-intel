"""
Create lease abstract files for S AND M Transportation / 330 Stevens Street
Updated: source citations added throughout PDF (page and section references)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

# ─────────────────────────────────────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────────────────────────────────────
OUTPUT_DIR = r"C:\Users\RhettAnderson\OneDrive - Entrada Management Services\Desktop\Claude\RIDGE\RIDGE Daily Reports"
XLSX_FILE = os.path.join(OUTPUT_DIR, "SandMTransportation_330Stevens_RentRoll_20260311.xlsx")
PDF_FILE  = os.path.join(OUTPUT_DIR, "SandMTransportation_330Stevens_LeaseAbstract_20260311.pdf")

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# STYLE CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
NAVY       = "1F3864"
YELLOW     = "FFFF00"
RED_FILL   = "FF0000"
ORANGE     = "FF6600"
LT_BLUE    = "BDD7EE"
GRAY       = "404040"
WHITE      = "FFFFFF"
LIGHT_GRAY = "D9D9D9"
NAVY_FILL2 = "2E4A78"

CURRENCY_FMT = '#,##0.00'
DOLLAR_FMT   = '"$"#,##0.00'


def navy_header_style():
    return {
        'fill': PatternFill("solid", fgColor=NAVY),
        'font': Font(bold=True, color=WHITE, size=12),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
    }


def apply_style(cell, fill=None, font=None, alignment=None, number_format=None):
    if fill:       cell.fill = fill
    if font:       cell.font = font
    if alignment:  cell.alignment = alignment
    if number_format: cell.number_format = number_format


def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)


def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — RENT SCHEDULE
# ─────────────────────────────────────────────────────────────────────────────
def build_rent_schedule(wb):
    ws = wb.create_sheet("Rent Schedule")

    headers = [
        "Period", "Start Date", "End Date",
        "Monthly Base Rent", "Monthly CAM (est.)",
        "Monthly Total", "Annual Base Rent", "$/SF/Year"
    ]

    # Write headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color=WHITE, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[1].height = 30

    # Data rows
    rows = [
        # Period, Start, End, MonthlyBase, MonthlyCam, MonthlyTotal, AnnualBase, PSF
        ("Abatement",          "04/01/2022", "06/30/2022",  0.00,      20031.25,  20031.25,   0.00,       0.00),
        ("Year 1A",            "07/01/2022", "08/31/2022",  51413.54,  20031.25,  71444.79,   None,       3.85),
        ("Year 1B",            "09/01/2022", "06/30/2023",  57422.92,  20031.25,  77454.17,   574229.20,  4.30),
        ("Year 2",             "07/01/2023", "06/30/2024",  59158.96,  "TBD reconciled", "—", 709907.52,  4.43),
        ("Year 3",             "07/01/2024", "06/30/2025",  60895.00,  "TBD",     "—",        730740.00,  4.56),
        ("Year 4 (CURRENT)",   "07/01/2025", "06/30/2026",  62764.58,  "TBD",     "—",        753174.96,  4.70),
        ("Year 5",             "07/01/2026", "07/31/2027",  64634.17,  "TBD",     "—",        775610.04,  4.84),
    ]

    # Year 1A annual note
    year1a_note = "(2 months)"

    for r_idx, row in enumerate(rows, 2):
        period, start, end, m_base, m_cam, m_total, ann_base, psf = row
        is_current = (period == "Year 4 (CURRENT)")

        # Determine fill
        if is_current:
            row_fill = PatternFill("solid", fgColor=YELLOW)
            row_font_bold = True
            font_color = "000000"
        else:
            row_fill = None
            row_font_bold = False
            font_color = "000000"

        def write(col, value, fmt=None):
            cell = ws.cell(row=r_idx, column=col, value=value)
            if row_fill:
                cell.fill = row_fill
            cell.font = Font(bold=row_font_bold, color=font_color)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border()
            if fmt:
                cell.number_format = fmt
            return cell

        write(1, period)
        write(2, start)
        write(3, end)

        # Monthly Base Rent
        if isinstance(m_base, float):
            write(4, m_base, DOLLAR_FMT)
        else:
            write(4, m_base)

        # Monthly CAM
        if isinstance(m_cam, float):
            write(5, m_cam, DOLLAR_FMT)
        else:
            write(5, m_cam)

        # Monthly Total
        if isinstance(m_total, float):
            write(6, m_total, DOLLAR_FMT)
        else:
            write(6, m_total)

        # Annual Base Rent
        if period == "Year 1A":
            write(7, year1a_note)
        elif isinstance(ann_base, float):
            write(7, ann_base, DOLLAR_FMT)
        else:
            write(7, ann_base)

        # $/SF/Year
        if isinstance(psf, float):
            write(8, psf, DOLLAR_FMT)
        else:
            write(8, psf)

    # Summary row
    summary_row = len(rows) + 2
    total_label_cell = ws.cell(row=summary_row, column=1, value="TOTAL BASE RENT (SUM)")
    total_label_cell.fill = PatternFill("solid", fgColor=NAVY)
    total_label_cell.font = Font(bold=True, color=WHITE)
    total_label_cell.alignment = Alignment(horizontal='center', vertical='center')
    total_label_cell.border = thin_border()

    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=6)

    total_val_cell = ws.cell(row=summary_row, column=7, value=3543661.72)
    total_val_cell.fill = PatternFill("solid", fgColor=NAVY)
    total_val_cell.font = Font(bold=True, color=WHITE)
    total_val_cell.number_format = DOLLAR_FMT
    total_val_cell.alignment = Alignment(horizontal='center', vertical='center')
    total_val_cell.border = thin_border()

    empty_cell = ws.cell(row=summary_row, column=8, value="")
    empty_cell.fill = PatternFill("solid", fgColor=NAVY)
    empty_cell.border = thin_border()

    # Note row
    note_row = summary_row + 1
    note_cell = ws.cell(row=note_row, column=1,
        value="Building SF: 160,250 | NNN Lease | CAM initial estimate $1.50/SF/year ($20,031.25/mo) with 5% annual cap")
    note_cell.font = Font(italic=True, size=10, color="404040")
    note_cell.alignment = Alignment(horizontal='left', wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)
    ws.row_dimensions[note_row].height = 20

    # Column widths
    col_widths = [22, 14, 14, 20, 22, 18, 20, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(w, 15)

    # Freeze top row
    ws.freeze_panes = "A2"

# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — CRITICAL DATES
# ─────────────────────────────────────────────────────────────────────────────
def build_critical_dates(wb):
    ws = wb.create_sheet("Critical Dates")

    headers = ["Date", "Event", "Priority", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color=WHITE, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[1].height = 30

    data = [
        ("04/01/2022", "Lease Commencement",             "Standard",  "63-month term begins"),
        ("04/01/2022", "First CAM Payment Due",           "Standard",  "$20,031.25/month"),
        ("08/2023",    "Lease Assignment Effective",      "HIGH",      "Miami Warehouse Logistics → S AND M Transportation, Inc."),
        ("07/01/2025", "Year 4 Rent Step",                "Standard",  "$62,764.58/month ($4.70/SF)"),
        ("07/01/2026", "Year 5 Rent Step",                "Standard",  "$64,634.17/month ($4.84/SF)"),
        ("10/29/2026", "RENEWAL OPTION NOTICE DEADLINE",  "CRITICAL",  "275 days prior to expiration; Tenant must notify or option expires"),
        ("07/31/2027", "Lease Expiration",                "HIGH",      "End of 63-month term; 16.5 months from today"),
        ("Within 90 days of year-end", "Annual CAM Reconciliation", "Standard", "Landlord sends statement; Tenant pays/receives credit within 30 days"),
        ("5th of each month", "Rent Late Threshold",      "Standard",  "Late charge = 5% if received after 5:00 PM Eastern"),
    ]

    for r_idx, (date, event, priority, notes) in enumerate(data, 2):
        if priority == "CRITICAL":
            row_fill = PatternFill("solid", fgColor=RED_FILL)
            row_font = Font(bold=True, color=WHITE)
        elif priority == "HIGH":
            row_fill = PatternFill("solid", fgColor=ORANGE)
            row_font = Font(bold=True, color=WHITE)
        else:
            # Alternating light blue / white
            if r_idx % 2 == 0:
                row_fill = PatternFill("solid", fgColor=LT_BLUE)
            else:
                row_fill = PatternFill("solid", fgColor=WHITE)
            row_font = Font(color="000000")

        for col, val in enumerate([date, event, priority, notes], 1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.fill = row_fill
            cell.font = row_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border()
        ws.row_dimensions[r_idx].height = 28

    # Column widths
    for col, w in enumerate([22, 36, 14, 55], 1):
        ws.column_dimensions[get_column_letter(col)].width = max(w, 15)

    ws.freeze_panes = "A2"

# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — LEASE SUMMARY
# ─────────────────────────────────────────────────────────────────────────────
def build_lease_summary(wb):
    ws = wb.create_sheet("Lease Summary")

    # Title row
    title_cell = ws.cell(row=1, column=1,
        value="LEASE SUMMARY — 330 STEVENS STREET, JACKSONVILLE FL")
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.font = Font(bold=True, color=WHITE, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 36

    # Section helper
    def section_header(row, label):
        for col in [1, 2]:
            c = ws.cell(row=row, column=col, value=label if col == 1 else "")
            c.fill = PatternFill("solid", fgColor=GRAY)
            c.font = Font(bold=True, color=WHITE, size=11)
            c.alignment = Alignment(horizontal='left', vertical='center')
            c.border = thin_border()
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 24

    def data_row(row, field, value, field_color=LT_BLUE, value_color=WHITE,
                 field_font=None, value_font=None):
        fc = ws.cell(row=row, column=1, value=field)
        vc = ws.cell(row=row, column=2, value=value)
        fc.fill = PatternFill("solid", fgColor=field_color)
        vc.fill = PatternFill("solid", fgColor=value_color)
        fc.font = field_font or Font(bold=True, color="000000")
        vc.font = value_font or Font(color="000000")
        fc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        vc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        fc.border = thin_border()
        vc.border = thin_border()
        ws.row_dimensions[row].height = 20

    current_row = 2

    # ── PROPERTY ──
    section_header(current_row, "PROPERTY"); current_row += 1
    property_data = [
        ("Address",       "330 Stevens Street, Jacksonville, FL 32254"),
        ("County",        "Duval County, State of Florida"),
        ("Building SF",   "160,250± SF"),
        ("Tenant's Share","100% (sole occupant)"),
        ("Use",           "Warehousing and Storage"),
    ]
    for f, v in property_data:
        data_row(current_row, f, v); current_row += 1

    # ── PARTIES ──
    section_header(current_row, "PARTIES"); current_row += 1
    parties_data = [
        ("Landlord",               "Milestone Jacksonville LLC (Delaware LLC)"),
        ("Landlord Address",       "5610 Wisconsin Ave, Suite G101, Chevy Chase, MD 20815"),
        ("Landlord Signatory",     "Scott Milestone, President & Owner"),
        ("Original Tenant",        "Miami Warehouse Logistics, Inc. (Florida corp)"),
        ("Current Tenant (Assignee)", "S AND M Transportation, Inc. (Alabama corp)"),
        ("Assignee Signatory",     "Ben Smith, GM"),
        ("Secondary Obligor",      "Miami Warehouse Logistics, Inc. (retains liability)"),
        ("Assignment Date",        "August 2023"),
    ]
    for f, v in parties_data:
        data_row(current_row, f, v); current_row += 1

    # ── LEASE TERM ──
    section_header(current_row, "LEASE TERM"); current_row += 1
    term_data = [
        ("Commencement",              "April 1, 2022"),
        ("Expiration",                "July 31, 2027"),
        ("Total Term",                "63 months"),
        ("Remaining Term (as of 3/11/26)", "~16.5 months"),
        ("Lease Type",                "NNN (Net Lease)"),
    ]
    for f, v in term_data:
        data_row(current_row, f, v); current_row += 1

    # ── FINANCIALS ──
    section_header(current_row, "FINANCIALS"); current_row += 1
    financials_data = [
        ("Current Monthly Base Rent",     "$62,764.58"),
        ("Current Annual Base Rent",      "$753,174.96"),
        ("Current Rent PSF (NNN)",        "$4.70/SF/year"),
        ("Final Period Monthly Rent",     "$64,634.17"),
        ("CAM Estimate (at commencement)","$20,031.25/month ($1.50/SF)"),
        ("CAM Escalation Cap",            "5% annually (controllable expenses)"),
        ("Management Fee in CAM",         "15% of Operating Expenses"),
        ("Security Deposit",              "$114,666.67 (2 months rent)"),
        ("Late Charge",                   "5% (if after 5th of month, 5:00 PM ET)"),
    ]
    for f, v in financials_data:
        data_row(current_row, f, v); current_row += 1

    # ── OPTION TO RENEW ──
    section_header(current_row, "OPTION TO RENEW"); current_row += 1
    option_data = [
        ("Number of Options",    "One (1) 5-year option"),
        ("Notice Deadline",      "October 29, 2026 (275 days prior)"),
        ("Renewal Rent",         "Greater of (a) last base rent + 5% + 3%/yr OR (b) FMV"),
        ("FMV Mechanism",        "Baseball appraisal arbitration; cap $2,500/party"),
        ("Commission on Renewal","None"),
    ]
    for f, v in option_data:
        data_row(current_row, f, v); current_row += 1

    # ── KEY PROVISIONS ──
    section_header(current_row, "KEY PROVISIONS"); current_row += 1
    provisions_data = [
        ("Assignment/Subletting",   "Prohibited without Landlord written consent"),
        ("First Right of Refusal",  "NONE (intentionally omitted)"),
        ("SNDA",                    "ABSENT — not included in lease"),
        ("Estoppel Certificate",    "Required within 10 days of Landlord notice"),
        ("Alterations",             "Prohibited without Landlord written consent"),
        ("Hazardous Materials",     "Prohibited without written consent; Tenant liability survives termination"),
        ("Default Notice (payment)","3 days"),
        ("Default Notice (other)",  "10 days"),
    ]
    for f, v in provisions_data:
        data_row(current_row, f, v); current_row += 1

    # ── MAINTENANCE RESPONSIBILITIES ──
    section_header(current_row, "MAINTENANCE RESPONSIBILITIES"); current_row += 1
    maint_data = [
        ("Landlord", "Roof, exterior walls, structural, underground plumbing, service lines to building"),
        ("Tenant",   "HVAC, interior, parking lot, landscaping, lighting, dock equipment, sprinklers, all utilities"),
    ]
    for f, v in maint_data:
        data_row(current_row, f, v); current_row += 1

    # ── RIDGE FLAGS ──
    section_header(current_row, "RIDGE FLAGS"); current_row += 1
    flags_data = [
        ("Near-term roll (16.5 months)",      "POSITIVE SOURCING SIGNAL — ownership motivation flag"),
        ("Below-market rent ($4.70/SF vs est. market $7-9/SF)", "Value-add opportunity"),
        ("No ROFR",                           "Clean acquisition — no tenant purchase right"),
        ("SNDA absent",                       "HIGH severity — confirm debt status before LOI"),
        ("Assignee credit quality",           "NEEDS VERIFICATION — S&M Transportation (Alabama, transportation sector)"),
        ("Renewal option notice",             "October 29, 2026 — acquisition timeline must account for this"),
        ("Gavel Risk Rating",                 "MEDIUM (watch for HIGH pending S&M credit check)"),
        ("RIDGE Conviction",                  "NEEDS MORE DATA"),
    ]
    high_severity = {"SNDA absent", "Assignee credit quality"}
    for f, v in flags_data:
        if f in high_severity:
            data_row(current_row, f, v,
                     field_color=NAVY,
                     value_color="FFD7D7",
                     field_font=Font(bold=True, color=WHITE),
                     value_font=Font(bold=True, color="CC0000"))
        else:
            data_row(current_row, f, v,
                     field_color="2E4A78",
                     value_color=WHITE,
                     field_font=Font(bold=True, color=WHITE))
        current_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 65

    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────────
# BUILD XLSX
# ─────────────────────────────────────────────────────────────────────────────
def build_xlsx():
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    build_rent_schedule(wb)
    build_critical_dates(wb)
    build_lease_summary(wb)

    wb.save(XLSX_FILE)
    print(f"[OK] Excel file saved: {XLSX_FILE}")


# ─────────────────────────────────────────────────────────────────────────────
# BUILD PDF  (with source citations)
# ─────────────────────────────────────────────────────────────────────────────
def build_pdf():
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        KeepTogether, HRFlowable, PageBreak
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    # ── Color palette ──
    NAVY_C    = colors.HexColor("#1F3864")
    YELLOW_C  = colors.HexColor("#FFFF00")
    RED_C     = colors.HexColor("#FF0000")
    ORANGE_C  = colors.HexColor("#FF6600")
    LT_BLUE_C = colors.HexColor("#BDD7EE")
    GRAY_C    = colors.HexColor("#404040")
    LT_GRAY_C = colors.HexColor("#F2F2F2")
    MED_GRAY_C= colors.HexColor("#D9D9D9")
    WHITE_C   = colors.white
    BLACK_C   = colors.black
    GREEN_C   = colors.HexColor("#375623")
    CITE_C    = colors.HexColor("#666666")   # citation gray

    PAGE_W, PAGE_H = letter
    MARGIN = 0.75 * inch

    doc = SimpleDocTemplate(
        PDF_FILE,
        pagesize=letter,
        leftMargin=MARGIN,
        rightMargin=MARGIN,
        topMargin=MARGIN,
        bottomMargin=MARGIN,
        title="Lease Abstract — 330 Stevens Street",
        author="RIDGE",
    )

    styles = getSampleStyleSheet()
    USABLE_W = PAGE_W - 2 * MARGIN

    # ── Paragraph style factory ──
    def ps(name, **kw):
        return ParagraphStyle(name, **kw)

    sTitle  = ps("sTitle",  fontSize=22, textColor=WHITE_C,  alignment=TA_CENTER,
                 fontName="Helvetica-Bold",    leading=28)
    sSub    = ps("sSub",    fontSize=11, textColor=WHITE_C,  alignment=TA_CENTER,
                 fontName="Helvetica",         leading=16)
    sSec    = ps("sSec",    fontSize=13, textColor=WHITE_C,  alignment=TA_LEFT,
                 fontName="Helvetica-Bold",    leading=18, leftIndent=6)
    sField  = ps("sField",  fontSize=9,  textColor=BLACK_C,
                 fontName="Helvetica-Bold",    leading=13)
    sVal    = ps("sVal",    fontSize=9,  textColor=BLACK_C,
                 fontName="Helvetica",         leading=13)
    sNote   = ps("sNote",   fontSize=8,  textColor=GRAY_C,
                 fontName="Helvetica-Oblique", leading=12)
    sFlag   = ps("sFlag",   fontSize=9,  textColor=RED_C,
                 fontName="Helvetica-Bold",    leading=13)
    sSmall  = ps("sSmall",  fontSize=8,  textColor=BLACK_C,
                 fontName="Helvetica",         leading=11)
    sSmallB = ps("sSmallB", fontSize=8,  textColor=BLACK_C,
                 fontName="Helvetica-Bold",    leading=11)
    sBig    = ps("sBig",    fontSize=14, textColor=NAVY_C,
                 fontName="Helvetica-Bold",    leading=20, alignment=TA_CENTER)
    sNormal = ps("sNormal", fontSize=9,  textColor=BLACK_C,
                 fontName="Helvetica",         leading=13)
    sAbsent = ps("sAbsent", fontSize=9,  textColor=RED_C,
                 fontName="Helvetica-Bold",    leading=13)
    # Citation style: 8pt, italic, gray
    sCite   = ps("sCite",   fontSize=8,  textColor=CITE_C,
                 fontName="Helvetica-Oblique", leading=10)
    # Source column header style
    sSrcHdr = ps("sSrcHdr", fontSize=8,  textColor=WHITE_C,
                 fontName="Helvetica-Bold",    leading=10, alignment=TA_CENTER)

    story = []

    # ── Helper: inline citation string (returns italic gray Paragraph) ──
    def cite(ref):
        """Return a small italic gray Paragraph for a source citation."""
        return Paragraph(ref, sCite)

    # ── Helper: value + citation stacked in one cell ──
    def val_with_cite(value_text, cite_text, val_style=None):
        """Combine value and citation into a single Paragraph using XML markup."""
        val_style = val_style or sVal
        # Build as two lines
        combined = (
            f'<para><font name="Helvetica" size="9">{value_text}</font><br/>'
            f'<font name="Helvetica-Oblique" size="7.5" color="#666666">{cite_text}</font></para>'
        )
        return Paragraph(combined, sVal)

    # ── Helper: section header ──
    def section_hdr(title):
        tbl = Table([[Paragraph(title, sSec)]], colWidths=[USABLE_W])
        tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), NAVY_C),
            ("TOPPADDING",    (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
            ("LEFTPADDING",   (0,0), (-1,-1), 8),
        ]))
        return tbl

    # ── Helper: three-col table with Source column ──
    # col widths: field | value | source
    def three_col_table(data_triples, col1_w=None, col2_w=None, col3_w=None):
        col1_w = col1_w or USABLE_W * 0.32
        col2_w = col2_w or USABLE_W * 0.50
        col3_w = col3_w or USABLE_W * 0.18
        rows = []
        for i, (f, v, src) in enumerate(data_triples):
            fp = Paragraph(f, sField) if isinstance(f, str) else f
            vp = Paragraph(v, sVal)   if isinstance(v, str) else v
            sp = Paragraph(src, sCite) if isinstance(src, str) else src
            rows.append([fp, vp, sp])

        tbl = Table(rows, colWidths=[col1_w, col2_w, col3_w], repeatRows=0)
        style_cmds = [
            ("GRID",          (0,0), (-1,-1), 0.5, MED_GRAY_C),
            ("TOPPADDING",    (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LEFTPADDING",   (0,0), (-1,-1), 6),
            ("RIGHTPADDING",  (0,0), (-1,-1), 6),
            ("VALIGN",        (0,0), (-1,-1), "TOP"),
        ]
        for i in range(0, len(rows), 2):
            style_cmds.append(("BACKGROUND", (0,i), (-1,i), LT_BLUE_C))
        for i in range(1, len(rows), 2):
            style_cmds.append(("BACKGROUND", (0,i), (-1,i), WHITE_C))

        tbl.setStyle(TableStyle(style_cmds))
        return tbl

    # ── Helper: three-col table with a header row ──
    def three_col_table_with_header(header_triple, data_triples,
                                    col1_w=None, col2_w=None, col3_w=None):
        col1_w = col1_w or USABLE_W * 0.32
        col2_w = col2_w or USABLE_W * 0.50
        col3_w = col3_w or USABLE_W * 0.18
        hdr_row = [Paragraph(h, sSrcHdr) for h in header_triple]
        rows = [hdr_row]
        for i, (f, v, src) in enumerate(data_triples):
            fp = Paragraph(f, sSmall) if isinstance(f, str) else f
            vp = Paragraph(v, sSmall) if isinstance(v, str) else v
            sp = Paragraph(src, sCite) if isinstance(src, str) else src
            rows.append([fp, vp, sp])

        tbl = Table(rows, colWidths=[col1_w, col2_w, col3_w], repeatRows=1)
        style_cmds = [
            ("BACKGROUND",    (0,0), (-1,0),  NAVY_C),
            ("TEXTCOLOR",     (0,0), (-1,0),  WHITE_C),
            ("GRID",          (0,0), (-1,-1), 0.5, MED_GRAY_C),
            ("TOPPADDING",    (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LEFTPADDING",   (0,0), (-1,-1), 5),
            ("VALIGN",        (0,0), (-1,-1), "TOP"),
        ]
        for i in range(1, len(rows), 2):
            style_cmds.append(("BACKGROUND", (0,i), (-1,i), LT_BLUE_C))
        for i in range(2, len(rows), 2):
            style_cmds.append(("BACKGROUND", (0,i), (-1,i), WHITE_C))
        tbl.setStyle(TableStyle(style_cmds))
        return tbl

    # ────────────────────────────────────
    # PAGE 1
    # ────────────────────────────────────

    # Title header box
    title_data = [
        [Paragraph("LEASE ABSTRACT", sTitle)],
        [Paragraph("330 Stevens Street  |  Jacksonville, FL 32254  |  160,250± SF", sSub)],
        [Paragraph("Prepared by RIDGE  |  March 11, 2026", sSub)],
    ]
    title_tbl = Table(title_data, colWidths=[USABLE_W])
    title_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), NAVY_C),
        ("TOPPADDING",    (0,0), (0,0),   16),
        ("BOTTOMPADDING", (0,2), (-1,-1), 14),
        ("TOPPADDING",    (0,1), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (0,0),   4),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
    ]))
    story.append(title_tbl)
    story.append(Spacer(1, 14))

    # ── PARTIES (with Source column) ──
    story.append(section_hdr("PARTIES  \u2014  Source: Pg. 1, Header; Pg. 16, Signatures; Pg. 17\u201320, Assignment Agreement"))
    story.append(Spacer(1, 2))
    story.append(three_col_table([
        ("Landlord",
         "Milestone Jacksonville LLC (Delaware LLC)",
         "Pg. 1, Header;\nPg. 16, Signatures"),
        ("Landlord Address",
         "5610 Wisconsin Ave, Suite G101, Chevy Chase, MD 20815",
         "Pg. 1, Header"),
        ("Landlord Signatory",
         "Scott Milestone, President & Owner",
         "Pg. 16, Signatures"),
        ("Original Tenant",
         "Miami Warehouse Logistics, Inc. (Florida corp)",
         "Pg. 1, Header"),
        ("Current Tenant (Assignee)",
         "S AND M Transportation, Inc. (Alabama corp)",
         "Pg. 17\u201320, Assignment Agmt."),
        ("Assignee Signatory",
         "Ben Smith, GM",
         "Pg. 17\u201320, Assignment Agmt."),
        ("Secondary Obligor",
         "Miami Warehouse Logistics, Inc. (retains liability post-assignment)",
         "Pg. 17\u201320, Assignment Agmt."),
        ("Assignment Date",
         "August 2023",
         "Pg. 17\u201320, Assignment Agmt."),
    ]))
    story.append(Spacer(1, 10))

    # ── LEASE TERM ──
    story.append(section_hdr("LEASE TERM"))
    story.append(Spacer(1, 2))
    story.append(three_col_table([
        ("Commencement Date",
         "April 1, 2022",
         "Pg. 1, \u00a73"),
        ("Expiration Date",
         "July 31, 2027",
         "Pg. 1, \u00a73"),
        ("Total Term",
         "63 months",
         "Pg. 1, \u00a73"),
        ("Remaining Term (as of 3/11/26)",
         "~16.5 months",
         "Pg. 1, \u00a73"),
        ("Lease Type",
         "NNN (Net Lease) \u2014 Tenant responsible for taxes, insurance, and maintenance",
         "Pg. 4, \u00a77"),
    ]))
    story.append(Spacer(1, 10))

    # ── BASE RENT SCHEDULE (with Source column) ──
    story.append(section_hdr("BASE RENT SCHEDULE  \u2014  Source: Pg. 1\u20132, \u00a74"))
    story.append(Spacer(1, 2))

    rent_headers = ["Period", "Start", "End", "Monthly Base", "Annual Base", "$/SF/Yr", "Source"]
    rent_rows_data = [
        ["Abatement",        "04/01/2022", "06/30/2022", "$0.00",        "$0.00",        "$0.00",  "Pg. 2, \u00a74(b)"],
        ["Year 1A",          "07/01/2022", "08/31/2022", "$51,413.54",   "(2 months)",   "$3.85",  "Pg. 1\u20132, \u00a74"],
        ["Year 1B",          "09/01/2022", "06/30/2023", "$57,422.92",   "$574,229.20",  "$4.30",  "Pg. 1\u20132, \u00a74"],
        ["Year 2",           "07/01/2023", "06/30/2024", "$59,158.96",   "$709,907.52",  "$4.43",  "Pg. 1\u20132, \u00a74"],
        ["Year 3",           "07/01/2024", "06/30/2025", "$60,895.00",   "$730,740.00",  "$4.56",  "Pg. 1\u20132, \u00a74"],
        ["Year 4 (CURRENT)", "07/01/2025", "06/30/2026", "$62,764.58",   "$753,174.96",  "$4.70",  "Pg. 1\u20132, \u00a74"],
        ["Year 5",           "07/01/2026", "07/31/2027", "$64,634.17",   "$775,610.04",  "$4.84",  "Pg. 1\u20132, \u00a74"],
        ["TOTAL",            "",           "",           "",              "$3,543,661.72","",        ""],
    ]

    # Column widths for 7-col rent table
    cw_base = USABLE_W / 7
    cw_src  = USABLE_W * 0.16
    cw_data = (USABLE_W - cw_src) / 6
    rent_col_widths = [cw_data] * 6 + [cw_src]

    rent_table_data = [[Paragraph(h, sSrcHdr if h == "Source" else sSmallB) for h in rent_headers]]
    for row in rent_rows_data:
        styled_row = []
        for i, c in enumerate(row):
            if i == 6:  # Source column
                styled_row.append(Paragraph(str(c), sCite))
            else:
                styled_row.append(Paragraph(str(c), sSmall))
        rent_table_data.append(styled_row)

    rent_tbl = Table(rent_table_data, colWidths=rent_col_widths, repeatRows=1)
    rent_style = [
        ("BACKGROUND",    (0,0),  (-1,0),  NAVY_C),
        ("TEXTCOLOR",     (0,0),  (-1,0),  WHITE_C),
        ("FONTNAME",      (0,0),  (-1,0),  "Helvetica-Bold"),
        ("GRID",          (0,0),  (-1,-1), 0.5, MED_GRAY_C),
        ("TOPPADDING",    (0,0),  (-1,-1), 4),
        ("BOTTOMPADDING", (0,0),  (-1,-1), 4),
        ("LEFTPADDING",   (0,0),  (-1,-1), 5),
        ("ALIGN",         (1,0),  (-1,-1), "CENTER"),
        # Alternating rows
        ("BACKGROUND",    (0,2),  (-1,2),  LT_BLUE_C),
        ("BACKGROUND",    (0,4),  (-1,4),  LT_BLUE_C),
        ("BACKGROUND",    (0,6),  (-1,6),  LT_BLUE_C),
        # Year 4 highlight (row index 6 in data = row 6 in table after header)
        ("BACKGROUND",    (0,6),  (-1,6),  YELLOW_C),
        ("FONTNAME",      (0,6),  (-1,6),  "Helvetica-Bold"),
        # Total row
        ("BACKGROUND",    (0,8),  (-1,8),  NAVY_C),
        ("TEXTCOLOR",     (0,8),  (-1,8),  WHITE_C),
        ("FONTNAME",      (0,8),  (-1,8),  "Helvetica-Bold"),
    ]
    rent_tbl.setStyle(TableStyle(rent_style))
    story.append(rent_tbl)
    story.append(Paragraph(
        "Building SF: 160,250  |  NNN Lease  |  CAM initial estimate $1.50/SF/yr ($20,031.25/mo) "
        "with 5% annual cap  |  Late charge: 5% if after 5th of month  (Pg. 1\u20132, \u00a74; Pg. 2, \u00a74(a))",
        sNote))
    story.append(Spacer(1, 10))

    # ── CAM & NNN CHARGES ──
    story.append(section_hdr("CAM & NNN CHARGES  \u2014  Source: Pg. 2\u20133, \u00a75"))
    story.append(Spacer(1, 2))
    story.append(three_col_table([
        ("CAM Estimate at Commencement",
         "$20,031.25/month ($1.50/SF/year)",
         "Pg. 2, \u00a74(c)"),
        ("CAM Escalation Cap",
         "5% annually (controllable expenses only)",
         "Pg. 2, \u00a75"),
        ("Management Fee",
         "15% of Operating Expenses (included in CAM)",
         "Pg. 3, \u00a75"),
        ("Operating Expenses Definition",
         "Defined per Pg. 3, \u00a75 — includes taxes, insurance, maintenance, management fee",
         "Pg. 3, \u00a75"),
        ("CAM Reconciliation",
         "Annual; Landlord delivers statement within 90 days of year-end; Tenant pays/credits within 30 days",
         "Pg. 2\u20133, \u00a75"),
        ("Tenant's Share",
         "100% (sole occupant of 160,250 SF building)",
         "Pg. 1, \u00a71"),
        ("NNN Components",
         "Tenant responsible for: property taxes, building insurance, all maintenance & repairs, utilities",
         "Pg. 4, \u00a77; Pg. 9, \u00a721"),
    ]))
    story.append(Spacer(1, 10))

    # ── SECURITY DEPOSIT ──
    story.append(section_hdr("SECURITY DEPOSIT  \u2014  Source: Pg. 3\u20134, \u00a76"))
    story.append(Spacer(1, 2))
    story.append(three_col_table([
        ("Amount",
         "$114,666.67 (equivalent to 2 months' base rent at execution)",
         "Pg. 3, \u00a76"),
        ("Return Timing",
         "Within 90 days of lease expiration, subject to Tenant compliance with all terms",
         "Pg. 4, \u00a76"),
        ("Late Charge",
         "5% of overdue amount if rent received after 5:00 PM Eastern on the 5th of the month",
         "Pg. 2, \u00a74(a)"),
    ]))
    story.append(Spacer(1, 18))

    # ────────────────────────────────────
    # PAGE 2
    # ────────────────────────────────────
    story.append(PageBreak())

    # ── OPTION TO RENEW ──
    story.append(section_hdr("OPTION TO RENEW  \u2014  Source: Pg. 15, \u00a741"))
    story.append(Spacer(1, 2))
    story.append(three_col_table([
        ("Number of Options",
         "One (1) renewal option",
         "Pg. 15, \u00a741"),
        ("Option Term",
         "5 years",
         "Pg. 15, \u00a741"),
        ("Notice Required",
         "275 days prior to expiration",
         "Pg. 15, \u00a741"),
        ("Notice Deadline",
         "October 29, 2026",
         "Pg. 15, \u00a741"),
        ("Renewal Rent",
         "Greater of: (a) final base rent + 5%, escalating 3%/year, OR (b) Fair Market Value",
         "Pg. 15, \u00a741"),
        ("FMV Mechanism",
         "Baseball appraisal arbitration; costs capped at $2,500 per party",
         "Pg. 15, \u00a741"),
        ("Commission on Renewal",
         "None",
         "Pg. 15, \u00a741"),
        ("Condition",
         "Tenant must not be in default at time of notice or commencement",
         "Pg. 15, \u00a741"),
    ]))
    story.append(Spacer(1, 6))

    # Callout box for renewal deadline
    renewal_callout_data = [[
        Paragraph("\u26a0  RENEWAL NOTICE DEADLINE: OCTOBER 29, 2026", ps("rc1",
            fontSize=11, textColor=WHITE_C, fontName="Helvetica-Bold",
            alignment=TA_CENTER, leading=16)),
    ],[
        Paragraph(
            "Tenant must deliver written notice of renewal election no later than October 29, 2026 "
            "(275 days prior to July 31, 2027 expiration).  Failure to timely notice results in "
            "irrevocable option expiration.  RIDGE acquisition timeline must account for this date.  "
            "<font name='Helvetica-Oblique' size='8' color='#FFDDDD'>(Pg. 15, \u00a741)</font>",
            ps("rc2", fontSize=9, textColor=WHITE_C, fontName="Helvetica",
               alignment=TA_CENTER, leading=13)),
    ]]
    renewal_callout = Table(renewal_callout_data, colWidths=[USABLE_W])
    renewal_callout.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), RED_C),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING",   (0,0), (-1,-1), 12),
        ("RIGHTPADDING",  (0,0), (-1,-1), 12),
    ]))
    story.append(renewal_callout)
    story.append(Spacer(1, 10))

    # ── PROVISION CHECKLIST (with Source column) ──
    story.append(section_hdr("PROVISION CHECKLIST"))
    story.append(Spacer(1, 2))

    prov_headers = ["Provision", "Status", "Notes", "Source"]
    prov_data = [
        ("Assignment / Subletting",   "PRESENT",  "Prohibited without prior written Landlord consent",          "Pg. 5\u20136, \u00a711"),
        ("First Right of Refusal",    "ABSENT",   "Intentionally omitted \u2014 clean acquisition",              "Pg. 14, \u00a740"),
        ("SNDA",                      "ABSENT",   "HIGH SEVERITY \u2014 verify lender status before LOI",        "Pg. 13, \u00a733"),
        ("Estoppel Certificate",      "PRESENT",  "Tenant must deliver within 10 days of Landlord notice",       "Pg. 14, \u00a738"),
        ("Option to Renew",           "PRESENT",  "One 5-year option; notice by 10/29/2026",                     "Pg. 15, \u00a741"),
        ("Renewal Commission",        "N/A",       "Lease states no commission on renewal",                       "Pg. 15, \u00a741"),
        ("Alterations",               "PRESENT",  "Prohibited without Landlord written consent",                 "Pg. 8, \u00a720"),
        ("Hazardous Materials",       "PRESENT",  "Prohibited; Tenant liability survives termination",           "Pg. 12, \u00a730"),
        ("CAM Audit Rights",          "PRESENT",  "Standard NNN reconciliation process",                         "Pg. 2\u20133, \u00a75"),
        ("Default / Cure Period",     "PRESENT",  "3 days (payment); 10 days (other defaults)",                  "Pg. 11, \u00a728"),
        ("Bankruptcy",                "PRESENT",  "Addressed in lease",                                          "Pg. 14, \u00a736"),
        ("Net Lease / No Setoff",     "PRESENT",  "Tenant may not withhold or offset rent",                      "Pg. 4, \u00a77"),
        ("Landlord Right of Entry",   "PRESENT",  "24-hour notice required",                                     "Pg. 7, \u00a715"),
        ("Subordination",             "PRESENT",  "No SNDA included",                                            "Pg. 13, \u00a733"),
    ]

    absent_items = {"ABSENT"}
    prov_cw_src = USABLE_W * 0.15
    prov_cw = [USABLE_W * 0.28, USABLE_W * 0.12, USABLE_W * 0.45, prov_cw_src]
    prov_table_rows = [[Paragraph(h, sSrcHdr if h == "Source" else sSmallB) for h in prov_headers]]
    for prov, status, notes, src in prov_data:
        status_style = sAbsent if status == "ABSENT" else sSmall
        prov_table_rows.append([
            Paragraph(prov, sSmall),
            Paragraph(status, status_style),
            Paragraph(notes, sSmall),
            Paragraph(src, sCite),
        ])

    prov_tbl = Table(prov_table_rows, colWidths=prov_cw, repeatRows=1)
    prov_style_cmds = [
        ("BACKGROUND",    (0,0),  (-1,0),  NAVY_C),
        ("TEXTCOLOR",     (0,0),  (-1,0),  WHITE_C),
        ("FONTNAME",      (0,0),  (-1,0),  "Helvetica-Bold"),
        ("GRID",          (0,0),  (-1,-1), 0.5, MED_GRAY_C),
        ("TOPPADDING",    (0,0),  (-1,-1), 4),
        ("BOTTOMPADDING", (0,0),  (-1,-1), 4),
        ("LEFTPADDING",   (0,0),  (-1,-1), 5),
        ("VALIGN",        (0,0),  (-1,-1), "TOP"),
    ]
    for i in range(1, len(prov_table_rows), 2):
        prov_style_cmds.append(("BACKGROUND", (0,i), (-1,i), LT_BLUE_C))
    prov_tbl.setStyle(TableStyle(prov_style_cmds))
    story.append(prov_tbl)
    story.append(Spacer(1, 10))

    # ── CRITICAL DATES (with Source column) ──
    story.append(section_hdr("CRITICAL DATES"))
    story.append(Spacer(1, 2))

    cd_headers = ["Date", "Event", "Priority", "Notes", "Source"]
    cd_data_rows = [
        ("04/01/2022", "Lease Commencement",             "Standard",  "63-month term begins",                                    "Pg. 1, \u00a73"),
        ("04/01/2022", "First CAM Payment Due",           "Standard",  "$20,031.25/month",                                        "Pg. 2, \u00a74(c)"),
        ("08/2023",    "Lease Assignment Effective",      "HIGH",      "Miami Warehouse Logistics \u2192 S AND M Transportation",  "Pg. 17\u201320, Assignment Agmt."),
        ("07/01/2025", "Year 4 Rent Step",                "Standard",  "$62,764.58/month ($4.70/SF)",                             "Pg. 1\u20132, \u00a74"),
        ("07/01/2026", "Year 5 Rent Step",                "Standard",  "$64,634.17/month ($4.84/SF)",                             "Pg. 1\u20132, \u00a74"),
        ("10/29/2026", "RENEWAL OPTION NOTICE DEADLINE",  "CRITICAL",  "275 days prior; failure = option expiration",             "Pg. 15, \u00a741"),
        ("07/31/2027", "Lease Expiration",                "HIGH",      "End of 63-month term; 16.5 months from today",            "Pg. 1, \u00a73"),
        ("Within 90 days YE", "Annual CAM Reconciliation","Standard",  "Landlord delivers statement; Tenant pays/credits in 30 days","Pg. 2\u20133, \u00a75"),
        ("5th of month", "Rent Late Threshold",           "Standard",  "Late charge = 5% if after 5:00 PM ET",                    "Pg. 2, \u00a74(a)"),
    ]

    cd_cw_src = USABLE_W * 0.16
    cd_cw = [USABLE_W * 0.15, USABLE_W * 0.25, USABLE_W * 0.11, USABLE_W * 0.33, cd_cw_src]
    cd_table_rows = [[Paragraph(h, sSrcHdr if h == "Source" else sSmallB) for h in cd_headers]]
    for date, event, priority, notes, src in cd_data_rows:
        row = [
            Paragraph(date, sSmall),
            Paragraph(event, sSmall),
            Paragraph(priority, sSmall),
            Paragraph(notes, sSmall),
            Paragraph(src, sCite),
        ]
        cd_table_rows.append(row)

    cd_tbl = Table(cd_table_rows, colWidths=cd_cw, repeatRows=1)
    cd_cmds = [
        ("BACKGROUND",    (0,0),  (-1,0),  NAVY_C),
        ("TEXTCOLOR",     (0,0),  (-1,0),  WHITE_C),
        ("FONTNAME",      (0,0),  (-1,0),  "Helvetica-Bold"),
        ("GRID",          (0,0),  (-1,-1), 0.5, MED_GRAY_C),
        ("TOPPADDING",    (0,0),  (-1,-1), 4),
        ("BOTTOMPADDING", (0,0),  (-1,-1), 4),
        ("LEFTPADDING",   (0,0),  (-1,-1), 5),
        ("VALIGN",        (0,0),  (-1,-1), "TOP"),
    ]
    # Alternate standard rows
    standard_rows = [i+1 for i, (_, _, p, _, _) in enumerate(cd_data_rows) if p == "Standard"]
    for i in standard_rows:
        cd_cmds.append(("BACKGROUND", (0,i), (-1,i), LT_BLUE_C if i % 2 == 0 else WHITE_C))
    # HIGH rows: orange
    high_rows = [i+1 for i, (_, _, p, _, _) in enumerate(cd_data_rows) if p == "HIGH"]
    for i in high_rows:
        cd_cmds.append(("BACKGROUND", (0,i), (-1,i), ORANGE_C))
        cd_cmds.append(("TEXTCOLOR",  (0,i), (-1,i), WHITE_C))
    # CRITICAL row: red
    critical_rows = [i+1 for i, (_, _, p, _, _) in enumerate(cd_data_rows) if p == "CRITICAL"]
    for i in critical_rows:
        cd_cmds.append(("BACKGROUND", (0,i), (-1,i), RED_C))
        cd_cmds.append(("TEXTCOLOR",  (0,i), (-1,i), WHITE_C))
        cd_cmds.append(("FONTNAME",   (0,i), (-1,i), "Helvetica-Bold"))

    cd_tbl.setStyle(TableStyle(cd_cmds))
    story.append(cd_tbl)
    story.append(Spacer(1, 18))

    # ────────────────────────────────────
    # PAGE 3
    # ────────────────────────────────────
    story.append(PageBreak())

    # ── RISK FLAGS (with Source column) ──
    story.append(section_hdr("RISK FLAGS"))
    story.append(Spacer(1, 6))

    def risk_block(label, color, items):
        """
        items: list of (item_label, item_text, source_ref) tuples.
        """
        label_cell = [[Paragraph(label, ps("rl", fontSize=12, textColor=WHITE_C,
            fontName="Helvetica-Bold", leading=16, alignment=TA_CENTER))]]
        label_tbl = Table(label_cell, colWidths=[USABLE_W])
        label_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), color),
            ("TOPPADDING",    (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ]))
        rows = []
        for item_label, item_text, src_ref in items:
            rows.append([
                Paragraph(item_label, ps("ri_l", fontSize=9, fontName="Helvetica-Bold",
                    textColor=BLACK_C, leading=13)),
                Paragraph(item_text, ps("ri_t", fontSize=9, fontName="Helvetica",
                    textColor=BLACK_C, leading=13)),
                Paragraph(src_ref, sCite),
            ])
        if rows:
            item_tbl = Table(rows, colWidths=[USABLE_W*0.26, USABLE_W*0.57, USABLE_W*0.17])
            item_tbl.setStyle(TableStyle([
                ("GRID",          (0,0), (-1,-1), 0.5, MED_GRAY_C),
                ("TOPPADDING",    (0,0), (-1,-1), 5),
                ("BOTTOMPADDING", (0,0), (-1,-1), 5),
                ("LEFTPADDING",   (0,0), (-1,-1), 6),
                ("VALIGN",        (0,0), (-1,-1), "TOP"),
                ("BACKGROUND",    (0,0), (-1,-1),
                    colors.HexColor("#FFF8F8") if color == RED_C
                    else (colors.HexColor("#FFF5EC") if color == ORANGE_C
                    else colors.HexColor("#F2F2F2"))),
            ]))
            return KeepTogether([label_tbl, item_tbl, Spacer(1, 8)])
        return KeepTogether([label_tbl, Spacer(1, 8)])

    story.append(risk_block("HIGH RISK ITEMS", RED_C, [
        ("SNDA Absent",
         "No SNDA in lease. If property is encumbered by debt, tenant's lease position is at risk in "
         "foreclosure. Must confirm existence/status of any lender and negotiate lender recognition "
         "agreement before proceeding to LOI. HIGH severity \u2014 potential deal-stopper.",
         "Pg. 13, \u00a733"),
        ("Assignee Credit Quality",
         "S AND M Transportation, Inc. is an Alabama transportation company. No credit verification "
         "performed. Transportation sector carries elevated default risk. Must obtain financials, D&B, "
         "or other credit indicators before underwriting rent income.",
         "Pg. 17\u201320, Assignment Agmt."),
    ]))

    story.append(risk_block("MEDIUM RISK ITEMS", ORANGE_C, [
        ("Near-Term Lease Expiration",
         "~16.5 months remaining (expiration 07/31/2027). Acquirer faces immediate leasing risk. "
         "However, this is the PRIMARY SOURCING SIGNAL \u2014 owner motivation likely elevated. "
         "Underwrite both lease renewal and vacancy scenarios.",
         "Pg. 1, \u00a73"),
        ("Renewal Option Notice Deadline",
         "October 29, 2026 \u2014 only ~7.5 months away. If acquired, new ownership must be prepared to "
         "engage tenant on renewal before this date. Renewal rent formula resets to market (FMV) "
         "which may significantly increase rent.",
         "Pg. 15, \u00a741"),
        ("Below-Market Rent",
         "$4.70/SF NNN vs. estimated Jacksonville industrial market of $7\u20139/SF NNN. "
         "Value-add upside on renewal/re-lease, but also confirms current owner has below-market "
         "income. Verify market comps before establishing FMV for renewal arbitration.",
         "Pg. 1\u20132, \u00a74"),
    ]))

    story.append(risk_block("LOW RISK / POSITIVE ITEMS", colors.HexColor("#375623"), [
        ("No ROFR",
         "First Right of Refusal intentionally omitted. Clean acquisition \u2014 no tenant has right to "
         "purchase the property and disrupt a sale.",
         "Pg. 14, \u00a740"),
        ("Original Tenant Remains Liable",
         "Miami Warehouse Logistics, Inc. retains secondary liability post-assignment. "
         "Provides additional credit backstop behind S&M Transportation.",
         "Pg. 17\u201320, Assignment Agmt."),
        ("100% Occupancy / Single Tenant",
         "Sole occupant leasing full 160,250 SF. Simplified management; no multi-tenant complexity.",
         "Pg. 1, \u00a71"),
        ("NNN Structure",
         "Full NNN lease. Tenant responsible for taxes, insurance, maintenance. "
         "Minimal landlord operating exposure during hold period.",
         "Pg. 4, \u00a77"),
    ]))

    story.append(Spacer(1, 10))

    # ── RIDGE CONVICTION ASSESSMENT ──
    story.append(section_hdr("RIDGE CONVICTION ASSESSMENT"))
    story.append(Spacer(1, 6))

    conviction_callout = Table([[
        Paragraph("NEEDS MORE DATA", ps("conv", fontSize=20, textColor=NAVY_C,
            fontName="Helvetica-Bold", alignment=TA_CENTER, leading=26))
    ]], colWidths=[USABLE_W])
    conviction_callout.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), YELLOW_C),
        ("TOPPADDING",    (0,0), (-1,-1), 14),
        ("BOTTOMPADDING", (0,0), (-1,-1), 14),
        ("BOX",           (0,0), (-1,-1), 2, NAVY_C),
    ]))
    story.append(conviction_callout)
    story.append(Spacer(1, 8))

    conviction_rationale = [
        ("Positive signals",
         "Near-term roll creates owner motivation. No ROFR. 100% NNN occupancy. Below-market rent "
         "creates value-add thesis if tenant renews or building is re-leased at market.",
         "Pg. 14, \u00a740; Pg. 4, \u00a77; Pg. 1, \u00a73"),
        ("Data gaps preventing conviction",
         "1) SNDA status / lender position unknown. "
         "2) S&M Transportation credit quality unverified. "
         "3) Market rent comps not formally sourced. "
         "4) Building condition / HVAC deferred maintenance unknown (Tenant-maintained). "
         "5) Current owner's basis and pricing expectations unknown.",
         "Pg. 13, \u00a733; Pg. 17\u201320"),
        ("Next steps to reach conviction",
         "Obtain: (a) SNDA or lender payoff confirmation; (b) S&M financials or D&B report; "
         "(c) 3\u20135 market rent comps for 100k+ SF Jacksonville industrial; "
         "(d) Informal pricing indication from broker/owner; "
         "(e) Basic building inspection summary if available.",
         "\u2014"),
    ]

    story.append(three_col_table(conviction_rationale,
                                 col1_w=USABLE_W*0.28,
                                 col2_w=USABLE_W*0.54,
                                 col3_w=USABLE_W*0.18))
    story.append(Spacer(1, 12))

    # ── GAVEL RISK RATING ──
    story.append(section_hdr("GAVEL RISK RATING"))
    story.append(Spacer(1, 6))

    gavel_header = Table([[
        Paragraph("MEDIUM", ps("gavel", fontSize=24, textColor=WHITE_C,
            fontName="Helvetica-Bold", alignment=TA_CENTER, leading=30))
    ]], colWidths=[USABLE_W])
    gavel_header.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), ORANGE_C),
        ("TOPPADDING",    (0,0), (-1,-1), 12),
        ("BOTTOMPADDING", (0,0), (-1,-1), 12),
    ]))
    story.append(gavel_header)
    story.append(Spacer(1, 6))

    gavel_explanation = Table([[Paragraph(
        "Current rating is MEDIUM, with potential escalation to HIGH pending S&M Transportation "
        "credit verification and SNDA / lender status confirmation.  The SNDA absence (Pg. 13, \u00a733) "
        "and unverified assignee credit quality (Pg. 17\u201320, Assignment Agmt.) are the two factors "
        "that could individually or collectively push the rating to HIGH.  The near-term expiration "
        "and below-market rent are priced-in risks that represent opportunity rather than threat, "
        "provided the market supports re-leasing or renewal at or above projected FMV "
        "(Pg. 15, \u00a741).  Do not proceed to LOI until SNDA status is confirmed.",
        sNormal)]], colWidths=[USABLE_W])
    gavel_explanation.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), LT_GRAY_C),
        ("BOX",           (0,0), (-1,-1), 1, ORANGE_C),
        ("TOPPADDING",    (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
        ("LEFTPADDING",   (0,0), (-1,-1), 12),
        ("RIGHTPADDING",  (0,0), (-1,-1), 12),
    ]))
    story.append(gavel_explanation)
    story.append(Spacer(1, 12))

    # Footer note
    story.append(HRFlowable(width=USABLE_W, thickness=0.5, color=GRAY_C))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        "This lease abstract was prepared by RIDGE for internal acquisitions analysis purposes only. "
        "It is based on a review of the lease document and assignment agreement and does not constitute "
        "legal advice. All financial figures should be independently verified prior to closing.  "
        "Source citations reference the original lease PDF page and section numbers.",
        ps("footer", fontSize=7, textColor=GRAY_C, fontName="Helvetica-Oblique",
           leading=10, alignment=TA_CENTER)))

    doc.build(story)
    print(f"[OK] PDF file saved: {PDF_FILE}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    build_xlsx()
    build_pdf()
    print("\nAll files created successfully.")
